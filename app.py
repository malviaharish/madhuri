import streamlit as st
import os
import time
import urllib.parse
import uuid
import zipfile
import shutil
import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# App layout & theme configuration (Must be the first Streamlit command)
st.set_page_config(
    page_title="Capture Studio — Search & Export Tool",
    page_icon="📸",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS Stylesheet for Premium Aesthetics
st.markdown("""
<style>
    /* Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&family=Inter:wght@400;500;600&display=swap');

    /* Global Typography & Font Overrides */
    html, body, [class*="css"], .stApp {
        font-family: 'Inter', sans-serif;
    }
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Outfit', sans-serif !important;
        font-weight: 600 !important;
        letter-spacing: -0.02em !important;
    }

    /* Glassmorphism sidebar */
    [data-testid="stSidebar"] {
        background-color: #0b0c10 !important;
        border-right: 1px solid rgba(255, 255, 255, 0.05);
    }

    /* Style container wrappers to look like glass cards */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 16px !important;
        border: 1px solid rgba(255, 255, 255, 0.06) !important;
        background-color: rgba(255, 255, 255, 0.015) !important;
        padding: 20px !important;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.2) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    [data-testid="stVerticalBlockBorderWrapper"]:hover {
        border-color: rgba(168, 85, 247, 0.25) !important;
        background-color: rgba(255, 255, 255, 0.025) !important;
        box-shadow: 0 12px 40px 0 rgba(0, 0, 0, 0.3) !important;
    }

    /* Custom Form & Input Styles */
    .stTextInput>div>div>input, .stSelectbox>div>div>div {
        border-radius: 10px !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        background-color: rgba(255, 255, 255, 0.02) !important;
        color: white !important;
        transition: all 0.25s ease !important;
    }
    .stTextInput>div>div>input:focus, .stSelectbox>div>div>div:focus-within {
        border-color: #a855f7 !important;
        box-shadow: 0 0 0 3px rgba(168, 85, 247, 0.2) !important;
    }

    /* Styling Playwright logs box */
    code, pre {
        background-color: #06070a !important;
        border: 1px solid rgba(255, 255, 255, 0.04) !important;
        border-radius: 12px !important;
        color: #2dd4bf !important; /* Cyber Teal */
        font-family: 'JetBrains Mono', monospace !important;
        font-size: 0.88rem !important;
        padding: 15px !important;
    }

    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px !important;
        background-color: transparent !important;
        border-bottom: 1px solid rgba(255, 255, 255, 0.05) !important;
        margin-bottom: 20px !important;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 20px !important;
        border-radius: 10px 10px 0 0 !important;
        background-color: rgba(255, 255, 255, 0.01) !important;
        border: 1px solid rgba(255, 255, 255, 0.05) !important;
        border-bottom: none !important;
        color: rgba(255, 255, 255, 0.5) !important;
        font-weight: 500 !important;
        transition: all 0.25s ease !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background-color: rgba(255, 255, 255, 0.04) !important;
        color: white !important;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.12) 0%, rgba(168, 85, 247, 0.12) 100%) !important;
        border-color: rgba(168, 85, 247, 0.3) !important;
        color: #e9d5ff !important;
        border-bottom: 2px solid #a855f7 !important;
    }

    /* Primary gradient button override */
    button[kind="primary"] {
        background: linear-gradient(135deg, #6366f1 0%, #a855f7 100%) !important;
        border: none !important;
        border-radius: 12px !important;
        color: white !important;
        font-weight: 600 !important;
        padding: 12px 28px !important;
        box-shadow: 0 4px 15px rgba(99, 102, 241, 0.25) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    button[kind="primary"]:hover {
        transform: translateY(-1.5px) !important;
        box-shadow: 0 6px 20px rgba(168, 85, 247, 0.45) !important;
    }

    /* Secondary buttons (Clear, deletes, etc.) */
    button[kind="secondary"] {
        border-radius: 10px !important;
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        background-color: rgba(255, 255, 255, 0.02) !important;
        transition: all 0.2s ease !important;
    }
    button[kind="secondary"]:hover {
        border-color: rgba(255, 255, 255, 0.15) !important;
        background-color: rgba(255, 255, 255, 0.04) !important;
    }

    /* File uploader container */
    [data-testid="stFileUploader"] {
        border: 2px dashed rgba(255, 255, 255, 0.08) !important;
        background: rgba(255, 255, 255, 0.005) !important;
        border-radius: 14px !important;
    }

    /* Image Gallery Cards */
    .gallery-img-container {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid rgba(255, 255, 255, 0.08);
        background: #0f172a;
        padding: 8px;
        transition: all 0.3s ease;
    }
    .gallery-img-container:hover {
        transform: translateY(-2px);
        border-color: rgba(168, 85, 247, 0.3);
    }
</style>
""", unsafe_allow_html=True)

# Auto-install Playwright browser if running in a cloud environment
try:
    import subprocess
    @st.cache_resource
    def install_playwright_browsers():
        subprocess.run(["playwright", "install", "chromium"], check=True)
    install_playwright_browsers()
except Exception as e:
    pass

# Initialize Session States
if "terms" not in st.session_state:
    st.session_state.terms = ["triclosan suture", "aspirin cardiovascular", "ibuprofen inflammation"]
if "logs" not in st.session_state:
    st.session_state.logs = []
if "outputs" not in st.session_state:
    st.session_state.outputs = []
if "job_id" not in st.session_state:
    st.session_state.job_id = None
if "is_running" not in st.session_state:
    st.session_state.is_running = False

# Workspace paths
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
JOBS_DIR = os.path.join(WORKSPACE_DIR, "jobs")
os.makedirs(JOBS_DIR, exist_ok=True)

def append_log(message, log_type="info"):
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    emoji = "ℹ️" if log_type == "info" else "✅" if log_type == "success" else "⚠️" if log_type == "warning" else "❌"
    log_entry = f"[{timestamp}] {emoji} {message}"
    st.session_state.logs.append(log_entry)
    print(log_entry)

def clean_page_content(page, engine):
    selectors = [
        "header", "footer", "#search-sidebar", ".ncbi-header", ".m-footer", 
        "#gb", "#footer", ".Appheader", "#sfdiv", "#b_header", "#b_footer", ".b_footer",
        "[id*='cookie']", "[class*='cookie']", "[id*='consent']", "[class*='consent']",
        "[id*='banner']", "[class*='banner']", ".cookie-banner", "#onetrust-consent-sdk"
    ]
    selector_str = ", ".join(selectors)
    js_cleanup = f"""
    try {{
        const elementsToHide = document.querySelectorAll("{selector_str}");
        elementsToHide.forEach(el => {{ if (el) el.style.display = 'none'; }});
        const mainContent = document.querySelector('main');
        if (mainContent) {{
            mainContent.style.width = '100%';
            mainContent.style.margin = '0';
            mainContent.style.padding = '20px';
        }}
        const containers = document.querySelectorAll('.container, #main, #content');
        containers.forEach(c => {{ if (c) {{ c.style.maxWidth = '100%'; c.style.width = '100%'; }} }});
    }} catch (e) {{}}
    """
    try:
        page.evaluate(js_cleanup)
    except Exception:
        pass

def get_search_url(engine, query, custom_template=None):
    encoded = urllib.parse.quote(query)
    if engine == 'pmc':
        return f"https://pmc.ncbi.nlm.nih.gov/search/?term={encoded}"
    elif engine == 'pubmed':
        return f"https://pubmed.ncbi.nlm.nih.gov/?term={encoded}"
    elif engine == 'google':
        return f"https://www.google.com/search?q={encoded}"
    elif engine == 'scholar':
        return f"https://scholar.google.com/scholar?q={encoded}"
    elif engine == 'custom' and custom_template:
        return custom_template.replace("{query}", encoded)
    return f"https://pmc.ncbi.nlm.nih.gov/search/?term={encoded}"

# Header Styling Redesigned
st.markdown("""
<div style="
    background: linear-gradient(135deg, rgba(15, 17, 26, 0.85) 0%, rgba(26, 27, 44, 0.85) 100%);
    backdrop-filter: blur(12px);
    border: 1px solid rgba(255, 255, 255, 0.08);
    padding: 30px;
    border-radius: 20px;
    margin-bottom: 30px;
    box-shadow: 0 10px 35px rgba(0, 0, 0, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.1);
    position: relative;
    overflow: hidden;
">
    <div style="
        position: absolute;
        top: -50px;
        right: -50px;
        width: 150px;
        height: 150px;
        background: radial-gradient(circle, rgba(168, 85, 247, 0.18) 0%, rgba(99, 102, 241, 0) 70%);
        border-radius: 50%;
        filter: blur(20px);
    "></div>
    <div style="display: flex; align-items: center; gap: 15px;">
        <span style="font-size: 2.6rem; filter: drop-shadow(0 4px 10px rgba(0,0,0,0.35));">📸</span>
        <div>
            <h1 style="
                margin: 0;
                font-size: 2.2rem;
                background: linear-gradient(135deg, #ffffff 40%, #c084fc 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                letter-spacing: -0.03em;
            ">Capture Studio</h1>
            <p style="margin: 4px 0 0 0; color: #94a3b8; font-size: 0.98rem; font-weight: 400;">
                Batch screenshot and PDF exporter for detailed web audits powered by <span style="color: #c084fc; font-weight: 500;">Playwright</span>.
            </p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)



# Layout Setup
col_left, col_right = st.columns([1, 2])

# Sidebar Settings
st.sidebar.title("⚙️ Control Settings")
engine = st.sidebar.selectbox("Search Engine", ["pmc", "pubmed", "google", "scholar", "custom"])
custom_url = ""
if engine == "custom":
    custom_url = st.sidebar.text_input("Custom URL Template", value="https://example.com/search?q={query}")

export_format = st.sidebar.selectbox("Export Format", ["screenshot", "screenshot_full", "pdf", "both"])
viewport = st.sidebar.selectbox("Viewport Size", ["1920x1080", "1280x720", "1536x864", "1024x768"])
clean_clutter = st.sidebar.checkbox("Remove Clutter (Header/Footer/Banners)", value=True)
dismiss_banners = st.sidebar.checkbox("Auto-Dismiss Cookie Banners", value=True)
delay = st.sidebar.slider("Render Delay (seconds)", min_value=1, max_value=10, value=3)

# --- Left Column: Search Terms Management ---
with col_left:
    st.subheader("📝 Search terms")
    
    # Excel Upload
    uploaded_file = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx"])
    if uploaded_file is not None:
        try:
            wb = load_workbook(uploaded_file, read_only=True)
            sheet = wb.active
            terms = []
            for row in sheet.iter_rows(values_only=True):
                if row and row[0] is not None:
                    val = str(row[0]).strip()
                    if val.lower() not in ["search term", "search terms", "query", "term", "queries", ""]:
                        terms.append(val)
            wb.close()
            if terms:
                st.session_state.terms = terms
                st.success(f"Loaded {len(terms)} terms from Excel!")
            else:
                st.warning("No search terms found in sheet.")
        except Exception as e:
            st.error(f"Error parsing Excel file: {e}")

    # Manual input
    new_term = st.text_input("Add manual term:")
    if st.button("➕ Add Term") and new_term.strip():
        st.session_state.terms.append(new_term.strip())
        st.rerun()

    # Terms List
    st.markdown("### Active Terms List")
    if st.session_state.terms:
        # Clear list button
        if st.button("🗑️ Clear All Terms"):
            st.session_state.terms = []
            st.rerun()

        for idx, term in enumerate(st.session_state.terms):
            col_t_text, col_t_del = st.columns([4, 1])
            col_t_text.markdown(f"**{idx + 1}.** `{term}`")
            if col_t_del.button("❌", key=f"del_{idx}"):
                st.session_state.terms.pop(idx)
                st.rerun()
    else:
        st.info("No active search terms. Type one above or upload an Excel sheet.")

# --- Right Column: Capture Dashboard ---
with col_right:
    st.subheader("🚀 Automation Dashboard")
    
    # Run Buttons
    col_run_1, col_run_2 = st.columns([1, 1])
    start_btn = col_run_1.button("▶️ Start Capture Job", disabled=st.session_state.is_running or not st.session_state.terms, use_container_width=True, type="primary")

    
    # Progress Indicators
    progress_bar = st.progress(0)
    progress_text = st.empty()
    
    # Log box container
    st.markdown("### 📋 Activity Logs")
    logs_container = st.empty()
    
    # Gallery Output Sections
    st.markdown("### 🖼️ Generated Outputs")
    output_tabs = st.tabs(["Screenshots", "PDF Documents"])
    
    # Run Job Process
    if start_btn:
        st.session_state.is_running = True
        st.session_state.logs = []
        st.session_state.outputs = []
        st.session_state.job_id = str(uuid.uuid4())
        job_dir = os.path.join(JOBS_DIR, st.session_state.job_id)
        os.makedirs(job_dir, exist_ok=True)
        
        append_log(f"Starting browser capture. Viewport: {viewport}, Engine: {engine}", "info")
        logs_container.code("\n".join(st.session_state.logs))
        
        width, height = map(int, viewport.split('x'))
        total_terms = len(st.session_state.terms)
        
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled"]
                )
                user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
                context = browser.new_context(
                    viewport={"width": width, "height": height},
                    user_agent=user_agent,
                    locale="en-US",
                    timezone_id="America/New_York",
                    extra_http_headers={
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                        "Accept-Language": "en-US,en;q=0.9",
                    }
                )
                
                # navigator footprint mask
                context.add_init_script("""
                    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                    window.navigator.chrome = { runtime: {}, loadTimes: () => {}, csi: () => {}, app: {} };
                """)
                
                page = context.new_page()
                
                for idx, term in enumerate(st.session_state.terms):
                    progress_percent = int(((idx) / total_terms) * 100)
                    progress_bar.progress(progress_percent)
                    progress_text.text(f"Processing term {idx + 1} of {total_terms}: '{term}'")
                    
                    append_log(f"[{idx+1}/{total_terms}] Navigating to: '{term}'", "info")
                    logs_container.code("\n".join(st.session_state.logs))
                    
                    url = get_search_url(engine, term, custom_url)
                    safe_chars = "".join([c if c.isalnum() or c in ".-_" else "_" for c in term]).strip("_")
                    
                    try:
                        page.goto(url, wait_until="domcontentloaded", timeout=25000)
                        time.sleep(delay)
                        
                        if dismiss_banners:
                            page.evaluate("""
                                const buttons = Array.from(document.querySelectorAll('button, a')).filter(el => {
                                    const t = el.innerText.toLowerCase();
                                    return t.includes('accept all') || t.includes('agree') || t.includes('accept cookies') || t.includes('i accept');
                                });
                                if (buttons.length > 0) buttons[0].click();
                            """)
                            time.sleep(0.5)
                        
                        if clean_clutter:
                            clean_page_content(page, engine)
                            time.sleep(0.5)
                        
                        outputs_generated = []
                        
                        # Screenshot Capture
                        if export_format in ["screenshot", "both"]:
                            filename = f"{idx+1}_screenshot_{safe_chars}.png"
                            filepath = os.path.join(job_dir, filename)
                            page.screenshot(path=filepath, full_page=False)
                            st.session_state.outputs.append({"type": "screenshot", "path": filepath, "name": filename, "term": term})
                            outputs_generated.append("Screenshot")
                            
                        # Full Page Screenshot Capture
                        if export_format == "screenshot_full":
                            filename = f"{idx+1}_screenshot_full_{safe_chars}.png"
                            filepath = os.path.join(job_dir, filename)
                            page.screenshot(path=filepath, full_page=True)
                            st.session_state.outputs.append({"type": "screenshot", "path": filepath, "name": filename, "term": term})
                            outputs_generated.append("Full Screenshot")

                        # PDF Capture
                        if export_format in ["pdf", "both"]:
                            filename = f"{idx+1}_document_{safe_chars}.pdf"
                            filepath = os.path.join(job_dir, filename)
                            page.pdf(path=filepath, format="A4", print_background=True)
                            st.session_state.outputs.append({"type": "pdf", "path": filepath, "name": filename, "term": term})
                            outputs_generated.append("PDF")
                            
                        append_log(f"✓ Successfully captured {', '.join(outputs_generated)} for '{term}'", "success")
                        
                    except Exception as page_err:
                        append_log(f"✗ Error processing '{term}': {page_err}", "error")
                    
                    logs_container.code("\n".join(st.session_state.logs))
                
                browser.close()
                
                # Bundle files into a ZIP
                zip_filename = f"export_job_{st.session_state.job_id}.zip"
                zip_path = os.path.join(job_dir, zip_filename)
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for out in st.session_state.outputs:
                        zipf.write(out["path"], out["name"])
                
                st.session_state.zip_path = zip_path
                st.session_state.zip_name = zip_filename
                
                progress_bar.progress(100)
                progress_text.text("Finished processing all search terms!")
                append_log("Job completed successfully. ZIP file is ready for download.", "success")
                logs_container.code("\n".join(st.session_state.logs))
                
        except Exception as job_err:
            append_log(f"Critical Automation Runner Error: {job_err}", "error")
            logs_container.code("\n".join(st.session_state.logs))
        
        st.session_state.is_running = False
        st.rerun()

    # Render Active Logs
    if st.session_state.logs:
        logs_container.code("\n".join(st.session_state.logs))

    # ZIP Download Button
    if "zip_path" in st.session_state and os.path.exists(st.session_state.zip_path):
        with open(st.session_state.zip_path, "rb") as f:
            st.download_button(
                label="📥 Download All Results (ZIP)",
                data=f,
                file_name=st.session_state.zip_name,
                mime="application/zip",
                use_container_width=True
            )

    # Render gallery items in tabs
    with output_tabs[0]:
        screenshots = [o for o in st.session_state.outputs if o["type"] == "screenshot"]
        if screenshots:
            # Display grid
            cols_gal = st.columns(3)
            for s_idx, s in enumerate(screenshots):
                col_gal = cols_gal[s_idx % 3]
                if os.path.exists(s["path"]):
                    col_gal.image(s["path"], caption=f"{s['term']}")
                    with open(s["path"], "rb") as sf:
                        col_gal.download_button(
                            label=f"Download Image",
                            data=sf,
                            file_name=s["name"],
                            mime="image/png",
                            key=f"dl_img_{s_idx}"
                        )
        else:
            st.info("No screenshots generated yet.")
            
    with output_tabs[1]:
        pdfs = [o for o in st.session_state.outputs if o["type"] == "pdf"]
        if pdfs:
            for p_idx, p in enumerate(pdfs):
                if os.path.exists(p["path"]):
                    col_pdf_name, col_pdf_dl = st.columns([4, 1])
                    col_pdf_name.markdown(f"📄 **{p['term']}** — `{p['name']}`")
                    with open(p["path"], "rb") as pf:
                        col_pdf_dl.download_button(
                            label="Download PDF",
                            data=pf,
                            file_name=p["name"],
                            mime="application/pdf",
                            key=f"dl_pdf_{p_idx}"
                        )
        else:
            st.info("No PDF files generated yet.")
